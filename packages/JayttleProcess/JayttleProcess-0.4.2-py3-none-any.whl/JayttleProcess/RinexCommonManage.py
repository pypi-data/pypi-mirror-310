import os
import shutil
import threading
import concurrent.futures
import subprocess
from enum import Enum
from collections import defaultdict, OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime, timedelta
from pathlib import Path
from time import sleep
from typing import Union, Optional
from JayttleProcess import CommonDecorator


class RinexFileType(Enum):
    O = 1
    N = 2
    Unkonw  = 3


class FileFormat(Enum):
    CRX = 1
    RNX = 2
    Unkonw  = 3


class StationType(Enum):
    RoverStation = 1
    BaseStation = 2
    Unkonw = 3


class RinexFileInfo:
    def __init__(self, input_path: str, ftp_file: bool):
        self.ftp_file: bool = ftp_file  # 是否来自FTP
        self.input_path: str = input_path  # 输入路径
        self.file_info: Path = Path(input_path)  # 文件信息对象
        self.file_name: str = self.file_info.name  # 文件名
        self.station_name: str = self.file_name[:3]  # 站点名称，前三个字符
        self.marker_name: str = self.file_name[:4]  # 标记名称，前四个字符
        self.station_id: int = int(self.station_name[2])  # 站点ID，从站点名称中第三个字符转换而来
        self.station_type: Union[StationType, str] = {  # 站点类型，根据站点名称的第一个字符确定，如果未知则为 Unknown
            'R': StationType.RoverStation,
            'B': StationType.BaseStation
        }.get(self.station_name[0], StationType.Unkonw)
        self.receiver_id: int = int(self.file_name[3])  # 接收器ID，从文件名第四个字符转换而来
        split = self.file_name.split('_')
        self.start_gps_time_str: str = split[2]  # 起始GPS时间字符串，从文件名中获取
        self.start_gps_time: datetime = self.get_time_from_string(self.start_gps_time_str)  # 起始GPS时间，使用自定义函数从字符串中获取
        self.duration_str: str = split[3]  # 持续时间字符串，从文件名中获取
        self.duration: timedelta = self.get_duration_from_string(self.duration_str)  # 持续时间，使用自定义函数从字符串中获取
        self.time_str: str = f"{self.start_gps_time_str}_{self.duration_str}"  # 时间字符串，包含起始GPS时间和持续时间
        file_type = split[-1][1]
        self.file_type: Union[RinexFileType, str] = {  # 文件类型，根据文件名中的第二个字符确定，如果未知则为 Unknown
            'O': RinexFileType.O,
            'N': RinexFileType.N
        }.get(file_type, RinexFileType.Unkonw)
        split1 = self.file_name.split('.')
        self.info_str: str = split1[0]  # 信息字符串，从文件名中获取
        compressed = split1[-1].lower()
        self.compressed: bool = compressed == "zip" or compressed == "z"  # 是否为压缩文件
        self.format: Union[FileFormat, str] = {    # 文件格式，根据文件扩展名确定，如果未知则为 Unknown
            "crx": FileFormat.CRX,
            "rnx": FileFormat.RNX
        }.get(split1[1].lower(), FileFormat.Unkonw)


    @staticmethod
    def get_time_from_string(time_str: str) -> datetime:
        """
        将字符串形式的时间转换为 datetime 对象。

        参数：
            time_str (str): 表示时间的字符串，格式为 'YYYYDDDHHMM'，其中：
                YYYY: 年份
                DDD: 年份中的第几天
                HH: 小时
                MM: 分钟

        返回：
            datetime: 表示转换后的时间的 datetime 对象。
        """
        year = int(time_str[:4])  # 年份
        day_of_year = int(time_str[4:7])  # 年份中的第几天
        hour = int(time_str[7:9])  # 小时
        minute = int(time_str[9:11])  # 分钟
        return datetime(year, 1, 1) + timedelta(days=day_of_year - 1, hours=hour, minutes=minute)


    @staticmethod
    def get_duration_from_string(duration_str: str) -> timedelta:
        """
        将字符串形式的时长转换为 timedelta 对象。

        参数：
            duration_str (str): 表示时长的字符串，格式为 'X[D/H/M/S]'，其中：
                X: 数字，表示时长的数量
                D: 天
                H: 小时
                M: 分钟
                S: 秒

        返回：
            timedelta: 表示转换后的时长的 timedelta 对象。
        """
        days = int(duration_str[:-1])
        unit = duration_str[-1]
        if unit == 'D':
            return timedelta(days=days)
        elif unit == 'H':
            return timedelta(hours=days)
        elif unit == 'M':
            return timedelta(minutes=days)
        elif unit == 'S':
            return timedelta(seconds=days)

    @staticmethod
    def get_string_from_duration(duration: timedelta) -> str:
        """
        将 timedelta 对象转换为字符串形式的时长。

        参数：
            duration (timedelta): 表示时长的 timedelta 对象。

        返回：
            str: 表示转换后的字符串形式的时长，格式为 '[X]D/H/M/S'，其中：
                X: 数字，表示时长的数量
                D: 天
                H: 小时
                M: 分钟
                S: 秒
        """
        if duration < timedelta(minutes=1):
            return f"{duration.total_seconds():.0f}S"
        elif duration < timedelta(hours=1):
            return f"{duration.total_seconds() / 60:.0f}M"
        elif duration < timedelta(days=1):
            return f"{duration.total_seconds() / 3600:.0f}H"
        else:
            return f"{duration.days}D"


def get_rnx_files(directory_path: str) -> list[RinexFileInfo]:
    """
    获取指定目录下的所有 Rinex 文件的信息。

    参数：
        directory_path (str): Rinex 文件所在目录的路径。

    返回：
        list[RinexFileInfo]: Rinex 文件信息的列表。
    """
    rnx_files = []
    for file_name in os.listdir(directory_path):
        if file_name.endswith(".rnx"):
            file_path = os.path.join(directory_path, file_name)
            rnx_file = RinexFileInfo(file_path, ftp_file=False)
            rnx_files.append(rnx_file)
    return rnx_files


class MergeFiles:
    @staticmethod
    @CommonDecorator.log_function_call
    def merge_files(files: list[RinexFileInfo], merge_path: str, rewrite: bool = False) -> bool:
        """
        合并多个 Rinex 文件。

        参数：
            files (list[RinexFileInfo]): 要合并的 Rinex 文件信息的列表。
            merge_path (str): 合并后文件存储的路径。
            rewrite (bool, 可选): 如果合并后的文件已存在，是否覆盖。默认为 False。

        返回：
            bool: 表示合并是否成功的布尔值。

        注意：
            如果文件列表为空或只有一个文件，则无法进行合并，会返回 False。
        """
        if files is None or len(files) <= 1:
            return False
        if not os.path.exists(merge_path):
            os.makedirs(merge_path)

        first = files[0]
        last = files[-1]

        split = first.file_name.split('_')
        start_time = first.start_gps_time
        start_time_str = first.start_gps_time_str
        duration = (last.start_gps_time - first.start_gps_time) + last.duration
        duration_str = RinexFileInfo.get_string_from_duration(duration)
        split[2] = start_time_str
        split[3] = duration_str
        merge_name = '_'.join(split)

        merge_path = os.path.join(merge_path, start_time_str)
        if not os.path.exists(merge_path):
            os.makedirs(merge_path)

        merge_full_name = os.path.join(merge_path, merge_name)  # Modify this line
        if os.path.exists(merge_full_name):
            if not rewrite:
                return True
            else:
                os.remove(merge_full_name)

        # 构建正确的参数列表
        file_names = [f.file_info for f in files]
        command = [r"D:\Program Files (x86)\Software\OneDrive\C#\WPF\Cableway.Download\Rinex\gfzrnx_2.1.0_win64.exe"]
        options = ["-finp"] + file_names + ["-kv", "-fout", merge_full_name, "-vo", "3.02"]

        # 执行外部程序
        process = subprocess.Popen(
            command + options,
            shell=False,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE
        )
        stdout, stderr = process.communicate()

        sleep(0.1)
        return os.path.exists(merge_full_name)


def filter_and_group_folders(directory: str) -> dict:
    """
    根据文件夹名称对文件夹进行过滤和分组。

    参数：
        directory (str): 要处理的文件夹路径。

    返回：
        dict: 包含分组信息的字典，键为日期的前八位数字，值为包含满足条件的小时数的列表。
    """
    folder_groups = {}

    for folder_name in os.listdir(directory):
        if folder_name[-2:].isdigit():
            last_two_digits = int(folder_name[-2:])
            first_eight_digits = int(folder_name[:8])

            if (last_two_digits >= 12 and last_two_digits <= 23):
                if first_eight_digits not in folder_groups:
                    folder_groups[first_eight_digits] = []
                folder_groups[first_eight_digits].append(last_two_digits)

    # 移除不满足条件的分组
    folder_groups = {key: value for key, value in folder_groups.items() if len(value) >= 4}

    return folder_groups


def group_by_hour(file_list: list[RinexFileInfo]) -> dict[str, list[RinexFileInfo]]:
    """
    将 RinexFileInfo 对象列表按照小时范围进行分组。

    参数：
        file_list (list[RinexFileInfo]): 包含 RinexFileInfo 对象的列表。

    返回：
        dict[str, list[RinexFileInfo]]: 按小时范围分组后的字典，键为小时范围字符串，值为对应的 RinexFileInfo 对象列表。
    """
    groups = {"12-15": [], "16-19": [], "20-23": []}
    for file_info in file_list:
        hour_key = int(file_info.start_gps_time.strftime("%H"))
        if 12 <= hour_key < 16:
            groups["12-15"].append(file_info)
        elif 16 <= hour_key < 20:
            groups["16-19"].append(file_info)
        elif 20 <= hour_key <= 23:
            groups["20-23"].append(file_info)
    return groups


def merge_files_by_hour(groups: dict[str, list[RinexFileInfo]], merge_path: str) -> None:
    """
    将按小时分组的 RinexFileInfo 对象列表合并到指定路径。

    参数：
        groups (dict[str, list[RinexFileInfo]]): 按小时范围分组的 RinexFileInfo 对象字典。
        merge_path (str): 合并后的文件存储路径。
    """
    for hour_range, file_group in groups.items():
        MergeFiles.merge_files(file_group, merge_path)


def merge_files_by_hour_multithread(groups: dict[str, list[RinexFileInfo]], merge_path: str) -> None:
    """
    使用多线程将按小时分组的 RinexFileInfo 对象列表合并到指定路径。

    参数：
        groups (dict[str, list[RinexFileInfo]]): 按小时范围分组的 RinexFileInfo 对象字典。
        merge_path (str): 合并后的文件存储路径。
    """
    threads = [threading.Thread(target=MergeFiles.merge_files, args=(file_group, merge_path)) for file_group in groups.values()]
    for thread in threads:
        thread.start()
    for thread in threads:
        thread.join()


def merge_files_by_hour_threadpool(groups: dict[str, list[RinexFileInfo]], merge_path: str) -> None:
    """
    使用线程池将按小时分组的 RinexFileInfo 对象列表合并到指定路径。

    参数：
        groups (dict[str, list[RinexFileInfo]]): 按小时范围分组的 RinexFileInfo 对象字典。
        merge_path (str): 合并后的文件存储路径。
    """
    # 使用 ThreadPoolExecutor 来管理线程池
    with concurrent.futures.ThreadPoolExecutor(max_workers=6) as executor:
        # 创建一个将 future 对象映射到小时范围的字典
        futures = {executor.submit(MergeFiles.merge_files, file_group, merge_path): hour_range for hour_range, file_group in groups.items()}
        # 遍历 future 对象，直到它们完成
        for future in concurrent.futures.as_completed(futures):
            try:
                # 尝试获取 future 的结果，如果出现异常则捕获并打印
                future.result()
            except Exception as e:
                print(f"发生异常：{e}")


def process_directory(directory_path: str, merge_path: str):
    """
    处理指定目录中的文件。

    参数：
    - directory_path：要处理的目录路径。
    - merge_path：合并文件后的保存路径。

    返回：
    无。
    """
    rnx_files: list[RinexFileInfo] = get_rnx_files(directory_path)
    o_files: list[RinexFileInfo] = [rnx_file for rnx_file in rnx_files if rnx_file.file_type == RinexFileType.O]
    n_files: list[RinexFileInfo] = [rnx_file for rnx_file in rnx_files if rnx_file.file_type == RinexFileType.N]
    
    # 按小时分组
    grouped_o_files: dict[str, list[RinexFileInfo]] = group_by_hour(o_files)
    grouped_n_files: dict[str, list[RinexFileInfo]] = group_by_hour(n_files)

    merge_files_by_hour_multithread(grouped_o_files, merge_path)
    merge_files_by_hour_multithread(grouped_n_files, merge_path)


def delete_directories(directories: list[str]):
    """
    删除指定的文件夹。

    参数：
    - directories：包含要删除的文件夹路径的列表。

    返回：
    无。
    """
    for directory in directories:
        if os.path.exists(directory):
            shutil.rmtree(directory)
            print(f"Deleted directory: {directory}")
        else:
            print(f"Directory does not exist: {directory}")


def proj_merge_rnx():
    directory_paths = [
        r"D:\Ropeway\GNSS\FTP\B011",
        r"D:\Ropeway\GNSS\FTP\B021",
        r"D:\Ropeway\GNSS\FTP\R031"
    ]
    merge_path = r"D:\Ropeway\GNSS\FTP\Merge"
    for directory_path in directory_paths:
        process_directory(directory_path, merge_path)

    # delete_directories(directory_paths)
    
# proj_merge_rnx()


def read_rinex_file_info(file_path: str) -> RinexFileInfo:
    """
    从 Rinex 文件名中解析 RinexFileInfo 对象。

    参数：
        file_path (str): Rinex 文件的路径。

    返回：
        RinexFileInfo: 解析得到的 Rinex 文件信息。
    """
    file_name = os.path.basename(file_path)
    file_name = file_name.replace(".crx.Z", ".rnx").replace(".rnx.Z", ".rnx")  # 统一替换文件类型后缀
    ftp_file = False  # 假设文件不是来自 FTP
    return RinexFileInfo(file_path, ftp_file)


def read_rinex_files_info(file_list_path: str) -> list[RinexFileInfo]:
    """
    从文件列表中逐行读取 Rinex 文件信息并返回 RinexFileInfo 对象列表。

    参数：
        file_list_path (str): 包含 Rinex 文件名的文本文件路径。

    返回：
        list[RinexFileInfo]: Rinex 文件信息的列表。
    """
    rnx_files_info = []
    with open(file_list_path, 'r') as file:
        for line in file:
            line = line.strip()  # 去除行末的换行符
            if line.endswith(".rnx") or line.endswith(".crx.Z") or line.endswith(".rnx.Z"):  # 确保是 Rinex 文件
                rnx_file_info = read_rinex_file_info(line)
                rnx_files_info.append(rnx_file_info)
    return rnx_files_info


def get_rnx_files_crx(directory_path: str) -> list[RinexFileInfo]:
    """
    获取指定目录下以 .crx 结尾的 Rinex 文件的信息列表。

    参数：
        directory_path (str): Rinex 文件所在目录的路径。

    返回：
        List[RinexFileInfo]: Rinex 文件信息的列表。
    """
    rinex_files_list = []
    for file_name in os.listdir(directory_path):
        if file_name.endswith(".crx"):
            crx_file_path = os.path.join(directory_path, file_name)
            rnx_file_path = os.path.join(directory_path, file_name.replace(".crx", ".rnx"))
            if not os.path.exists(rnx_file_path):
                rinex_file = RinexFileInfo(crx_file_path, ftp_file=False)
                rinex_files_list.append(rinex_file)
    return rinex_files_list


def get_rnx_files_dict(directory_path: str) -> dict[datetime, list[RinexFileInfo]]:
    """
    获取指定目录下的所有 Rinex 文件的信息，并以起始GPS时间为键，RinexFileInfo 对象为值的列表形式返回。

    参数：
        directory_path (str): Rinex 文件所在目录的路径。

    返回：
        Dict[datetime, List[RinexFileInfo]]: Rinex 文件信息的字典，以起始GPS时间为键，对应的 RinexFileInfo 对象列表为值。
    """
    rinex_files_dict = {}
    for file_name in os.listdir(directory_path):
        if file_name.endswith(".rnx") or file_name.endswith(".crx"):
            file_path = os.path.join(directory_path, file_name)
            rinex_file = RinexFileInfo(file_path, ftp_file=False)
            start_gps_time = rinex_file.start_gps_time
            if start_gps_time not in rinex_files_dict:
                rinex_files_dict[start_gps_time] = []
            rinex_files_dict[start_gps_time].append(rinex_file)
    return rinex_files_dict


def get_rnx_files_dict_date(directory_path: str) -> dict[datetime.date, list[RinexFileInfo]]:
    """
    获取指定目录下的所有 Rinex 文件的信息，并以日期为键，RinexFileInfo 对象为值的列表形式返回。

    参数：
        directory_path (str): Rinex 文件所在目录的路径。

    返回：
        Dict[date, List[RinexFileInfo]]: Rinex 文件信息的字典，以日期为键，对应的 RinexFileInfo 对象列表为值。
    """
    rinex_files_dict = {}
    for file_name in os.listdir(directory_path):
        if file_name.endswith(".rnx") or file_name.endswith(".crx"):
            file_path = os.path.join(directory_path, file_name)
            rinex_file = RinexFileInfo(file_path, ftp_file=False)
            start_gps_time = rinex_file.start_gps_time
            start_date = start_gps_time.date()
            if start_date not in rinex_files_dict:
                rinex_files_dict[start_date] = []
            rinex_files_dict[start_date].append(rinex_file)
    return rinex_files_dict


def create_marker_name_excel(marker_name_files_per_day_1h: dict[datetime.date, dict], marker_name: str, file_path: str) -> None:
    # 创建一个新的工作簿
    wb = Workbook()

    # 写入每天中持续时间为 01 小时的文件数量，包括标题行
    ws_1h = wb.active
    ws_1h.title = "Files per Day (01H)"
    ws_1h['A1'] = "Date"
    ws_1h['B1'] = "File Count"
    ws_1h['C1'] = "Percentage"

    # 设置标题行样式
    for cell in ['A1', 'B1', 'C1']:
        ws_1h[cell].alignment = Alignment(horizontal='center')

    # 获取当前 marker_name 的最大文件数量
    max_count = max(marker_name_files_per_day_1h[marker_name].values())

    # 遍历对应 marker_name 的数据，写入到工作表中
    for idx, (date, count) in enumerate(marker_name_files_per_day_1h[marker_name].items(), start=2):
        ws_1h[f'A{idx}'] = date.strftime('%Y-%m-%d')
        ws_1h[f'B{idx}'] = count
        # 计算百分比
        percentage = count / max_count * 100 if max_count != 0 else 0
        ws_1h[f'C{idx}'] = f"{percentage:.2f}%"

    # 保存工作簿
    wb.save(file_path)


def Proj_export_excel(rinex_files_info: list[RinexFileInfo]):
    # 创建一个嵌套的 defaultdict 来存储不同 marker_name 的每天中持续时间为 01 小时的文件数量
    marker_name_files_per_day_1h = defaultdict(lambda: defaultdict(int))
    # 遍历 RinexFileInfo 列表
    for file_info in rinex_files_info:
        # 获取 marker_name
        marker_name = file_info.marker_name
        
        # 获取持续时间字符串
        duration_str = file_info.duration_str

        # 检查持续时间是否为 01H
        if duration_str[-3:] == "01H":
            # 获取文件起始时间的日期
            start_date: datetime.date = file_info.start_gps_time.date()

            # 增加对应 marker_name 和日期的文件数量
            marker_name_files_per_day_1h[marker_name][start_date] += 1

    # 输出到 Excel 表格中
    for marker_name in marker_name_files_per_day_1h.keys():
        file_path = f"{marker_name}_statistics.xlsx"  # 根据 marker_name 构造文件名
        create_marker_name_excel(marker_name_files_per_day_1h, marker_name, file_path)


def create_excel_file(files_per_day_1h: dict[str, int], duration_str_counts: dict[str, int]) -> None:
    # 创建一个新的工作簿
    wb = Workbook()

    # 写入每天中持续时间为 01 小时的文件数量，包括标题行
    ws_1h = wb.create_sheet(title="Files per Day (01H)")
    ws_1h['A1'] = "Date"
    ws_1h['B1'] = "File Count"

    # 设置标题行样式
    for cell in ['A1', 'B1']:
        ws_1h[cell].alignment = Alignment(horizontal='center')

    for idx, (date, count) in enumerate(files_per_day_1h.items(), start=2):
        ws_1h[f'A{idx}'] = date.strftime('%Y-%m-%d')
        ws_1h[f'B{idx}'] = count

    # 设置列宽和单元格对齐方式
    for col in ['A', 'B']:
        ws_1h.column_dimensions[col].width = 13.88
        for row in ws_1h.iter_rows(min_row=2, max_row=len(files_per_day_1h)+1, min_col=ws_1h[col][0].column, max_col=ws_1h[col][0].column):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')

    # 写入持续时间字符串的种类及对应的文件数量，包括标题行
    ws_duration = wb.create_sheet(title="Duration Counts")
    ws_duration['A1'] = "Duration Type"
    ws_duration['B1'] = "File Count"

    # 设置标题行样式
    for cell in ['A1', 'B1']:
        ws_duration[cell].alignment = Alignment(horizontal='center')

    for idx, (duration_str, count) in enumerate(duration_str_counts.items(), start=2):
        ws_duration[f'A{idx}'] = duration_str
        ws_duration[f'B{idx}'] = count

    # 设置列宽和单元格对齐方式
    for col in ['A', 'B']:
        ws_duration.column_dimensions[col].width = 13.88
        for row in ws_duration.iter_rows(min_row=2, max_row=len(duration_str_counts)+1, min_col=ws_duration[col][0].column, max_col=ws_duration[col][0].column):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')

    # 删除默认的工作表
    default_sheet = wb['Sheet']
    wb.remove(default_sheet)

    # 保存工作簿
    excel_file_path = "statistics.xlsx"
    wb.save(excel_file_path)


def count_files_per_day(rinex_files_info: list[RinexFileInfo]) -> dict[datetime.date, dict[str, int]]:
    """
    统计每天文件数，并按日期和标记站名返回文件数的字典。

    参数：
        rinex_files_info (list[RinexFileInfo]): Rinex 文件信息的列表。

    返回：
        OrderedDict[date, OrderedDict[str, int]]: 每天文件数的有序字典，键为日期，值为另一个有序字典，
                                                   其中键为标记站名，值为文件数。
    """
    # 初始化一个包含所有标记站名的列表
    all_marker_names = ['R031', 'R032', 'R051', 'R052', 'R071', 'R072', 'R081', 'R082', 'B011', 'B021']

    files_per_day = defaultdict(lambda: defaultdict(int))
    
    for rinex_file in rinex_files_info:
        file_date = rinex_file.start_gps_time.date()  # 获取文件的日期信息
        marker_name = rinex_file.marker_name  # 获取标记站名
        
        # 获取持续时间字符串
        duration_str = rinex_file.duration_str

        # 检查持续时间是否为 01H
        if duration_str[-3:] == "01H":
            # 在字典中增加或更新日期和标记站名对应的文件数
            files_per_day[file_date][marker_name] += 1
    
    # 确保每个日期都包含所有的标记站名，将缺失的文件数填充为0
    for files in files_per_day.values():
        for marker_name in all_marker_names:
            files[marker_name] = files.get(marker_name, 0)
    
    # 对结果进行排序
    sorted_files_per_day = OrderedDict(sorted(files_per_day.items()))
    for files in sorted_files_per_day.values():
        sorted_files = OrderedDict(sorted(files.items()))
        files.clear()
        files.update(sorted_files)
    
    return sorted_files_per_day


def count_files_in_hour_range(rinex_files_info: list[RinexFileInfo], start_hour: int, end_hour: int) -> dict[datetime.date, dict[str, int]]:
    """
    统计每天指定时间范围内文件的数量，并按日期和标记站名返回文件数的字典。

    参数：
        rinex_files_info (list[RinexFileInfo]): Rinex 文件信息的列表。
        start_hour (int): 起始小时。
        end_hour (int): 结束小时。

    返回：
        OrderedDict[date, OrderedDict[str, int]]: 每天文件数的有序字典，键为日期，值为另一个有序字典，
                                                   其中键为标记站名，值为文件数。
    """
    # 初始化一个包含所有标记站名的列表
    all_marker_names = ['R031', 'R032', 'R051', 'R052', 'R071', 'R072', 'R081', 'R082', 'B011', 'B021']

    files_per_day = defaultdict(lambda: defaultdict(int))
    
    for rinex_file in rinex_files_info:
        file_date = rinex_file.start_gps_time.date()  # 获取文件的日期信息
        file_hour = rinex_file.start_gps_time.hour  # 获取文件的小时信息
        marker_name = rinex_file.marker_name  # 获取标记站名
        

        # 获取持续时间字符串
        duration_str = rinex_file.duration_str

        # 检查持续时间是否为 01H
        if duration_str[-3:] == "01H":
            # 检查文件的小时是否在指定范围内
            if start_hour <= file_hour <= end_hour:
                # 在字典中增加或更新日期和标记站名对应的文件数
                files_per_day[file_date][marker_name] += 1
    
    # 确保每个日期都包含所有的标记站名，将缺失的文件数填充为0
    for files in files_per_day.values():
        for marker_name in all_marker_names:
            files[marker_name] = files.get(marker_name, 0)
    
    # 对结果进行排序
    sorted_files_per_day = OrderedDict(sorted(files_per_day.items()))
    for files in sorted_files_per_day.values():
        sorted_files = OrderedDict(sorted(files.items()))
        files.clear()
        files.update(sorted_files)
    
    return sorted_files_per_day


def find_dates_with_specific_file_count(files_in_hour_range: dict[datetime.date, dict[str, int]], marker_names: list[str], count: int) -> list[datetime.date]:
    """
    查找指定标记站名文件数同时为特定数量的日期列表。

    参数：
        files_in_hour_range (dict[datetime.date, dict[str, int]]): 包含文件数信息的字典。
        marker_names (list[str]): 要检查的标记站名列表。
        count (int): 要检查的文件数。

    返回：
        list[datetime.date]: 符合条件的日期列表。
    """
    dates_with_specific_count = []

    for date, file_counts in files_in_hour_range.items():
        # 检查每个标记站名的文件数是否满足要求
        if all(file_counts.get(marker_name, 0) == count for marker_name in marker_names):
            dates_with_specific_count.append(date)

    return dates_with_specific_count


def creat_excel(export_data: dict[datetime.date, dict[str, int]]) -> None:
    # 创建一个新的Excel工作簿
    wb = Workbook()
    ws = wb.active

    # 写入表头
    header = ["日期", "B011", "B021", "R031", "R032", "R051", "R052", "R071", "R072", "R081", "R082"]
    ws.append(header)

    # 将数据写入表格
    for date, file_counts in export_data.items():
        row = [date.strftime("%Y-%m-%d")] + [file_counts.get(marker, 0) for marker in header[1:]]
        ws.append(row)

    # 保存Excel文件
    output_excel_path = "D:\\Ropeway\\GNSS\\files_in_hour_range.xlsx"
    wb.save(output_excel_path)


def write_rinex_files_to_txt(rinex_files_info: list[RinexFileInfo], 
                              marker_name: str, 
                              dates: list[datetime.date], 
                              output_file_path: str,
                              start_hour: int,
                              end_hour: int) -> None:
    with open(output_file_path, 'w') as output_file:
        for rinex_file_info in rinex_files_info:
            file_hour = rinex_file_info.start_gps_time.hour
            file_date = rinex_file_info.start_gps_time.date()
            if rinex_file_info.marker_name == marker_name and file_date in dates and start_hour <= file_hour <= end_hour:
                output_file.write(f"{rinex_file_info.file_name}\n")



def extract_with_winrar(file_path: str) -> None:
    """
    使用 WinRAR 解压指定的 .Z 文件。

    参数：
        file_path (str): 要解压的 .Z 文件的完整路径。
    """
    # 构建 WinRAR 命令
    winrar_command = f'WinRAR x "{file_path}"'

    # 调用 WinRAR 解压文件
    subprocess.run(winrar_command, shell=True)


def unzip_file(file_path: str) -> bool:
    """解压单个文件并返回是否成功"""
    try:
        # 使用 WinRAR 解压文件（在后台执行）
        cmd = f'"C:\\Program Files\\WinRAR\\WinRAR.exe" x -y "{file_path}" "{os.path.dirname(file_path)}"'
        process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell=True)
        _, error_output = process.communicate()

        # 检查解压是否成功
        if process.returncode == 0:
            print(f"Successfully extracted {os.path.basename(file_path)}")
            return True
        else:
            print(f"Extraction failed for {os.path.basename(file_path)}: {error_output.decode()}")
            return False
    except Exception as e:
        print(f"Error occurred while extracting {os.path.basename(file_path)}: {e}")
        return False

def unzip_folder_path(folder_path: str = r"D:\Ropeway\GNSS\FTP\B021") -> None:
    """使用WinRAR工具解压缩指定文件夹的文件"""
    # 存储解压任务的列表
    unzip_tasks = []

    # 遍历文件夹中的所有文件
    for file_name in os.listdir(folder_path):
        # 检查文件是否以 .Z 结尾
        if file_name.endswith(".Z"):
            # 构建文件的完整路径
            file_path = os.path.join(folder_path, file_name)
            # 添加解压任务到列表
            unzip_tasks.append(file_path)

    # 使用 ThreadPoolExecutor 创建线程池
    with concurrent.futures.ThreadPoolExecutor() as executor:
        # 提交解压任务给线程池
        results = executor.map(unzip_file, unzip_tasks)

    # 删除所有成功解压的 .Z 文件
    for file_path, result in zip(unzip_tasks, results):
        if result:  # 如果解压成功
            os.remove(file_path)


def crx_to_rnx(rinex_files: list[RinexFileInfo], delete_crx_file: bool = True, overwrite: bool = True) -> None:
    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
        futures = {executor.submit(convert_file, rnx_info, delete_crx_file): rnx_info for rnx_info in rinex_files}
        for future in concurrent.futures.as_completed(futures):
            try:
                future.result()
            except Exception as e:
                print(f"发生异常：{e}")
                

def convert_file(rnx_info: RinexFileInfo, delete_crx_file: bool) -> None:
    if rnx_info.format == FileFormat.CRX:
        process = subprocess.Popen([r"C:\Program Files\Trimble\Trimble Business Center\CRX2RNX.exe",
                                    rnx_info.file_info],
                                    stdout=subprocess.PIPE,
                                    stderr=subprocess.PIPE)
        process.wait()
        if process.returncode == 0:
            if delete_crx_file and os.path.exists(rnx_info.file_info):
                os.remove(rnx_info.file_info)
        else:
            print("CRX to RNX conversion failed with error:", process.stderr.read())




def merge_files_threadpool(file_groups: list[list[RinexFileInfo]], merge_path: str, merge_file_num: int) -> None:
    """
    使用线程池将 RinexFileInfo 对象列表合并到指定路径。

    参数：
        file_groups (list[list[RinexFileInfo]]): RinexFileInfo 对象列表的列表。
        merge_path (str): 合并后的文件存储路径。
        merge_file_num (int): 每个 RinexFileInfo 对象列表中所需的文件数量。
    """
    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
        for rinex_files in file_groups:
            if len(rinex_files) != merge_file_num:
                print(f"文件数量不符合要求，跳过合并操作：{rinex_files}")
                continue
            future = executor.submit(MergeFiles.merge_files, rinex_files, merge_path)
            try:
                future.result()
            except Exception as e:
                print(f"发生异常：{e}")


def process_file(file_path: str, file_name: str) -> None:
    """
    处理单个文件的函数
    """
    if int(file_name[16:19]) > 216:
        print(f"{file_name}超日期跳过")
        return
    # 读取文件内容
    with open(file_path, 'r', encoding='utf-8') as file:
        # 读取文件的所有行数据
        lines = file.readlines()

    # 获取文件名的前四个字符
    file_name_prefix = file_name[:4]
    print(file_name_prefix)
    # 遍历文件的前四行数据，查找以 "MARKER NAME" 结尾的行并替换该行的前四个字符为文件名的前四个字符
    for i in range(min(4, len(lines))):
        if lines[i].strip().endswith("MARKER NAME"):
            # 如果行内容已经是要替换的内容，则直接结束替换过程
            if lines[i].startswith(file_name_prefix):
                print(f"{file_name}已完成")
                return
            
            # 构造新的行内容
            new_line = file_name_prefix + lines[i][4:]
            # 将新的行内容替换原来的行
            lines[i] = new_line
            print(f"{file_name}第{i+1}行替换为{file_name_prefix}")
            break
    else:
        print("前四行数据中未找到以 'MARKER NAME' 结尾的行。")
        return

    # 将修改后的内容写回文件
    with open(file_path, 'w', encoding='utf-8') as file:
        file.writelines(lines)



def process_rnx_files(folder_path: str) -> None:
    """
    对每个文件夹中末尾是 _MO.rnx 的文件执行指定操作，将第四行数据的前四个字符修改为文件名的前四个字符。
    在修改之前检查第四行数据的前四个字符是否已经等于文件名的前四个字符。如果是，则跳过该文件
    """
    # 获取目标文件夹中所有文件夹的名字
    subdirectories = os.listdir(folder_path)
    # 存储符合条件的文件名的字典，键是目录路径，值是文件名列表
    rnx_files = {}
    
    for directory in subdirectories:
        # 构建文件夹的完整路径
        directory_path = os.path.join(folder_path, directory)
        # 检查路径是否是文件夹
        if os.path.isdir(directory_path):
            # 初始化当前目录的文件名列表
            rnx_files[directory_path] = []
            # 遍历文件夹中的文件
            for file_name in os.listdir(directory_path):
                # 检查文件是否是以 "_MO.rnx" 结尾的文件
                if file_name.endswith("_MO.rnx"):
                    if int(file_name[16:19]) < 216 and int(file_name[12:16]) == 2023:
                        # 添加符合条件的文件名到文件名列表
                        rnx_files[directory_path].append(file_name)

    # 对字典中每个目录的文件名列表进行排序
    for directory, files in rnx_files.items():
        files.sort()

    # 遍历字典，处理每个目录中的文件
    for directory, files in rnx_files.items():
        for file in files:
            file_path = os.path.join(directory, file)
            process_file(file_path=file_path,file_name=file)

def count_lines_in_each_txt_file(folder_path: str) -> dict[str, int]:
    file_lines_mapping = {}
    for root, dirs, files in os.walk(folder_path):
        for file_name in files:
            if file_name.endswith('.txt'):
                file_path = os.path.join(root, file_name)
                with open(file_path, 'r', encoding='utf-8') as file:
                    lines = file.readlines()
                    num_lines = len(lines)
                    file_lines_mapping[file_name[:4]] = num_lines
    return file_lines_mapping


def count_files_in_folders(root_folder: str) -> dict:
    folder_file_count = {}
    small_file_count = 0  # 记录文件大小小于 1 KB 的个数
    for root, dirs, files in os.walk(root_folder):
        folder_name = os.path.relpath(root, root_folder)  # 获取相对于根文件夹的文件夹名称
        crx_count = sum(1 for file in files if file.endswith('.crx'))  # 统计当前文件夹下的 crx 文件数量
        rnx_count = sum(1 for file in files if file.endswith('.rnx'))  # 统计当前文件夹下的 rnx 文件数量
        for file in files:
            file_path = os.path.join(root, file)
            file_size_kb = os.path.getsize(file_path) / 1024  # 将文件大小转换为 KB
            if file_size_kb < 1:  # 如果文件大小小于 1 KB
                small_file_count += 1
        folder_file_count[folder_name] = {'crx': crx_count, 'rnx': rnx_count, 'small_files': small_file_count}  # 将文件夹名称及其文件数量添加到字典中
    return folder_file_count

def count_files_in_folders_version2(root_folder: str) -> dict:
    start_time_count = {}  # 在循环外部定义 start_time_count
    for root, _, files in os.walk(root_folder):
        directory_path = os.path.relpath(root, root_folder)
        for file_name in files:
            if file_name.endswith(".rnx"):
                file_path = os.path.join(directory_path, file_name)
                rnx_file = RinexFileInfo(file_path, ftp_file=False)
                start_time = rnx_file.start_gps_time_str
                if start_time in start_time_count:
                    start_time_count[start_time].append(file_name)
                else:
                    start_time_count[start_time] = [file_name]
        
    # 检查是否存在相同的 start_gps_time_str
    for start_time, files in start_time_count.items():
        if len(files) < 6:
            print(f"For start_gps_time_str: {start_time}, there is <6 file ")


def delete_small_files(directory_path: str, threshold_size: float) -> None:
    for file_name in os.listdir(directory_path):
        file_path = os.path.join(directory_path, file_name)
        file_size_bytes = os.path.getsize(file_path)
        file_size_kb = file_size_bytes / 1024  # 将文件大小转换为 KB
        if file_size_kb < threshold_size:
            os.remove(file_path)  # 删除文件

if __name__ == '__main__':
    process_file(r"D:\Ropeway\FTPMerge\20230651600\B02100000_R_20230651600_4H_01S_MO.rnx", f"R03100000_R_20232171200_4H_01S_MO.rnx")