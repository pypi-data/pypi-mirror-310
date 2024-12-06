"""reporting module

This module uses an ical loader package to fill a pandas database which is parsed to generate a workbook and worksheets

The difficulty may be to handle recurrent events.

note..:
    existing ical loading packages
    - ixsts (local, non pip) from N. Garcia-Rosa
    - icalendar
    - icalevents (fork of icalendar)
    - ics-py
    - ical (handle recurring events)
"""
import re
from datetime import datetime, timedelta
from ical.calendar_stream import IcsCalendarStream
import pandas as pd
from pathlib import Path
import openpyxl as xl
from openpyxl.utils.dataframe import dataframe_to_rows

_default_startdate = "2020-01-01"
_default_enddate = "2030-01-01"

def ical_to_dframe(filename: Path, startdate: datetime, enddate: datetime):
    """read a file in ICAL format and create events between 2 dates
    can handle recurrent events

    Args:
        filename (Path): path name of ICAL file
        startdate (datetime): starting date for events
        enddate (datetime): ending date (included) for events

    Returns:
        DataFrame: as member of the class (date, year, month, h_begin, duration, name, description, begin) properties
    """
    with filename.open() as ics_file:
        calendar = IcsCalendarStream.calendar_from_ics(ics_file.read())
    return pd.DataFrame(
        [
            {
                "date": event.dtstart,
                "year": event.dtstart.year,
                "month": event.dtstart.month,
                "h_begin": event.dtstart.time,
                "duration": (event.dtend - event.dtstart).seconds / 3600.0,
                "name": event.summary,
                "description": event.description,
                "begin": event.dtstart,
            }
            for event in calendar.timeline.included(startdate, enddate)
        ]
    )


class IcalList():
    _ObjName = "Ical List"

    def __init__(self, name: str, folder=None, start: str = _default_startdate, end: str = _default_enddate):
        self._name = name
        self._folder = name if folder is None else folder
        self._start = datetime.fromisoformat(start)
        self._end = datetime.fromisoformat(end) + timedelta(days=1)
        print(f"> init {self._ObjName} {self._name} in folder {self._folder}")
        print(f"    will include {self._start.date()} to {self._end.date()} (not included)")

    def load_ics(self):
        """load ICS files and create pandas dataframe associated to member"""
        framedict = {}
        filelist = list(Path(self._folder).glob("*.ics"))
        for filename in filelist:
            print(f"- reading {filename}")
            framedict[filename.stem] = ical_to_dframe(filename, self._start, self._end)
            framedict[filename.stem]["Member"] = filename.stem  # file name without path nor extension
        if len(filelist) == 0:
            raise FileNotFoundError("no *.ics file found")
        self._dframe = pd.concat(tuple(framedict.values()))
        self._autoparse()
        self.clean()

    def _autoparse(self):
        pass

    def clean(self):
        pass

    def filter(self, start: str, end: str):
        """replace DataFrame with selected dates"""
        self._dframe = self._df_slot(start, end)

    def members(self):
        """returns set (unordered list) of members"""
        return set(self._dframe["Member"])


class Project(IcalList):
    """Class Project, can read ICAL files, define tags and create worksheet 
    """

    _ObjName = "Project"
    _WP_model = r"WP[0-9]*[A-Z]*"

    def __init__(self, name: str, folder=None, start: str = _default_startdate, end: str = _default_enddate, default_WP="No_WP"):
        super().__init__(name, folder, start, end)
        self._default_WP = default_WP

    def _autoparse(self):
        self._set_wp(default=self._default_WP)

    def clean(self):
        self._clean_wp()

    def work_packages(self):
        """returns set (unordered list) of Work Packages"""
        return set(self._dframe["WP"])

    def _df_to_tab(self, wb: xl.Workbook, df, title=None):
        ws = wb.create_sheet(title=title)
        for row in dataframe_to_rows(df, header=True):
            ws.append(row)
        return ws

    def _df_slot(self, start: str, end: str):
        df = self._dframe
        slot = df[df["date"] >= start]
        slot = slot[slot["date"] <= end]
        return slot

    def _set_wp(self, default=None):
        """set WP properties using regexp in event name"""
        rewp = re.compile(self._WP_model)
        self._dframe["WP"] = self._dframe["name"].apply(lambda s: rewp.findall(s))
        self._dframe["WP"] = self._dframe["WP"].apply(lambda wplist: wplist[0] if wplist else default)

    def _clean_wp(self):
        rewp = re.compile(self._WP_model + r". *-* *")
        self._dframe["name"] = self._dframe["name"].apply(lambda s: rewp.sub("", s))
        return

    def add_tabdetail_member(self, wb, member: str):
        df = self._dframe[self._dframe["Member"] == member].loc[:, ["date", "duration", "WP", "name"]]
        df["date"] = df["date"].apply(lambda d: d.strftime("%d/%m/%Y"))
        ws = self._df_to_tab(wb, df, member)
        ws.delete_cols(1)
        ws.delete_rows(2)
        return ws

    def add_tab_workpackage(self, wb, wp: str):
        dfwp = self._dframe[self._dframe["WP"] == wp].loc[
            :, ["date", "Member", "duration", "WP", "name", "year", "month"]
        ]
        monthformat = lambda m : f"{m:02}"
        dfwp["YearMonth"] = dfwp.year.map(str) + "-" + dfwp.month.map(monthformat)
        dfpiv = dfwp.pivot_table(
            values="duration", index="Member", columns="YearMonth", aggfunc="sum"
        ).fillna(0)
        ws = self._df_to_tab(wb, dfpiv, wp)
        ws.delete_rows(2)
        return ws

    def add_tab_allworkpackages(self, wb):
        df = self._dframe.pivot_table(values="duration", index="Member", columns="WP", aggfunc="sum")
        ws = self._df_to_tab(wb, df, "Synthèse")
        ws.delete_rows(2)

    def workbook(self) -> xl.Workbook:
        print("> create workbook")
        wb = xl.Workbook()
        ws_empty = wb.active
        print("- create global worksheet")
        self.add_tab_allworkpackages(wb)
        for member in self.members():
            print(f"- create member worksheet {member}")
            self.add_tabdetail_member(wb, member)
        for wp in sorted(self.work_packages()):
            if wp is not None:
                print(f"- create WP worksheet {wp}")
                self.add_tab_workpackage(wb, wp)
        wb.remove(ws_empty)
        return wb

class Hyperplanning(IcalList):
    """Class Hyperplanning, can read ICAL files, define tags and create worksheet 
    """

    _ObjName = "Hyperplanning"
    _course_model = r"^[0-9a-zA-Z-_]* - "

    def __init__(self, name: str, folder=None, start: str = _default_startdate, end: str = _default_enddate, default_course="No_course"):
        super().__init__(name, folder, start, end)
        self._default_course = default_course

    def _autoparse(self):
        self._set_course()
        self._set_type()

    def clean(self):
        self._clean_course()
        self._clean_type()

    def courses(self):
        """returns set (unordered list) of Work Packages"""
        return set(self._dframe["Course"])

    def _df_to_tab(self, wb: xl.Workbook, df, title=None):
        ws = wb.create_sheet(title=title)
        for row in dataframe_to_rows(df, header=True):
            ws.append(row)
        return ws

    def _df_slot(self, start: str, end: str):
        df = self._dframe
        slot = df[df["date"] >= start]
        slot = slot[slot["date"] <= end]
        return slot

    def _set_course(self, default="unknown"):
        """set Course properties using regexp in event name"""
        rewp = re.compile(self._course_model)
        self._dframe["Course"] = self._dframe["name"].apply(lambda s: rewp.findall(s))
        self._dframe["Course"] = self._dframe["Course"].apply(lambda wplist: wplist[0][:-3] if wplist else default)

    def _clean_course(self):
        rewp = re.compile(self._course_model)
        self._dframe["name"] = self._dframe["name"].apply(lambda s: rewp.sub("", s))
        return

    def _set_type(self, default="unknown"):
        """set course type (Type) properties using regexp in event name"""
        rewp = re.compile(" - [a-zA-Z\\*]*$")
        self._dframe["CourseType"] = self._dframe["name"].apply(lambda s: rewp.findall(s))
        self._dframe["CourseType"] = self._dframe["CourseType"].apply(lambda wplist: wplist[0][3:] if wplist else default)

    def _clean_type(self):
        rewp = re.compile(" - [a-zA-Z\\*]*$")
        self._dframe["name"] = self._dframe["name"].apply(lambda s: rewp.sub("", s))
        return

    def add_tabdetail_member(self, wb, member: str):
        df = self._dframe[self._dframe["Member"] == member].loc[:, ["date", "duration", "Course", "CourseType", "name"]]
        df["date"] = df["date"].apply(lambda d: d.strftime("%d/%m/%Y"))
        ws = self._df_to_tab(wb, df, member)
        ws.delete_cols(1)
        ws.delete_rows(2)
        return ws

    def add_tab_course(self, wb, course: str):
        dfwp = self._dframe[self._dframe["Course"] == course].loc[
            :, ["date", "Member", "duration", "Course", "TypeCourse", "name", "year", "month"]
        ]
        dfpiv = dfwp.pivot_table(
            values="duration", index="Member", columns="CourseType", aggfunc="sum"
        ).fillna(0)
        ws = self._df_to_tab(wb, dfpiv, course)
        ws.delete_rows(2)
        return ws

    def add_tab_allcourses(self, wb):
        df = self._dframe.pivot_table(values="duration", index="CourseType", columns="Course", aggfunc="sum")
        ws = self._df_to_tab(wb, df, "Synthèse")
        ws.delete_rows(2)

    def workbook(self) -> xl.Workbook:
        print("> create workbook")
        wb = xl.Workbook()
        ws_empty = wb.active
        print("- create global worksheet")
        self.add_tab_allcourses(wb)
        for member in self.members():
            print(f"- create member worksheet {member}")
            self.add_tabdetail_member(wb, member)
            #self.add_tabsumup_member(wb, member)
        # for course in sorted(self.courses()):
        #     if course is not None:
        #         print(f"- create course worksheet {wp}")
        #         self.add_tab_course(wb, wp)
        wb.remove(ws_empty)
        return wb


if __name__ == "__main__":
    prj = Project(name="mambo", folder="examples/projectA", start="2023-01-01", end="2024-01-01")
    prj.load_ics()  # lecture des .ics
    # prj.filter(start="2023-01-01", end="2023-12-31") # filtre des dates
    wb = prj.workbook()  # création du tableur
    wb.save("projetA.xlsx")  # sauvegarde du fichier
