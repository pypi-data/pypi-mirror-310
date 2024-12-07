
import gzip
import csv
import itertools
import operator
import zipfile
import threading
import time
import queue
import sys 
import inspect

from dateutil.parser import parse

from .utils.sqltab import get_sqltab_count

__all__ = ['source_sql','source_pandas','source_csv','source_unionall']

source_unionall = itertools.chain

class Source(object):
    def __rshift__(self, other):
        from datamation import table_basic,Dimension,FactTable
        if isinstance(other,(table_basic,Dimension,FactTable)):
            self.tag_out = other
            self.execute = operator.methodcaller('insert')
        else:
            raise TypeError

        if self.tag_out:
            self.startup = self._startup
             
    def __and__(self,other):
        '''union operation Source & Source
        '''
        return self.unionall(other)
        
    def _startup(self):
        for row in self:
            self.execute(self.tag_out,row)
        self.tag_out.endload()

    def unionall(self,other):
        if issubclass(other.__class__,self):
           self.__iter__ = itertools.chain(self,other)
           return self

class source_sql(Source):
    def __init__(self,conn,sql,params=None,arraysize =20000,rename={},data_format = 'dict'):
        '''数据库查询
        Parameters
        --------------
        conn: PEP249 API
            数据库连接
        sql: str
            sql语句
        arraysize:int
            每次请求数据库批量返回行数 根据性能合理配置
        Returns
        ---------
        dict
            返回一个每行数据的迭代器
        '''
        self.rowcount = get_sqltab_count(conn,sql)
        # 原始连接
        self.__conn = conn
        # 可使用连接
        self.connect = conn() if callable(conn) else conn
        self.cursor = self.connect.cursor()
        self.cursor.arraysize = arraysize
        self.itersize = 5000
        self.data_format = data_format
        if params:
            self.cursor.execute(sql,params)
        else:
            self.cursor.execute(sql)
        self.cols = [col[0] for col in self.cursor.description]
        self.cols = [rename.get(n,n) for n in self.cols]

    def __iter__(self):
        return self.__next__()

    def __next__(self):
        while 1:
            data = self.cursor.fetchone()
            if data:
                if self.data_format=='dict':
                    if isinstance(data,dict):
                        yield data
                    else:
                        yield dict(zip(self.cols, data))
                else:
                    yield data
            else:
                if callable(self.__conn):
                    self.connect.close()
                break
            
    def __len__(self):
        return self.rowcount

class source_pandas(Source):
    def __init__(self, dataFrame,rename={}):
        '''pandas数据源迭代器 {columns:values}格式
        '''
        import pandas as pd
        self._dataFrame = dataFrame.rename(columns = rename)
        self._dataFrame = self._dataFrame.where(self._dataFrame.notnull(), None)
        self.cols = [n for n in dataFrame.columns]
    
    def __iter__(self):
        return self.__next__()

    def __next__(self):
        for row in self._dataFrame.to_dict('records'): 
            yield row

class source_xlsx(Source):
    def __init__(self,file,sheet_name='',rename={},to_datetime = []):
        '''xlsx数据源
        '''
        from openpyxl import load_workbook
        self.wb = load_workbook(filename = file,data_only=True)
        if sheet_name:
            self.ws = self.wb[sheet_name]
        else:
            self.ws = self.wb[self.wb.sheetnames[0]]
        self.cols = [str(n).lower() for n in self.ws.iter_rows(min_row=1, max_col=self.ws.max_column, max_row=1,values_only=True)]
        rename = {str(k).lower():v for k,v in rename.items()}
        self.cols = [n if n not in rename else rename[n] for n in self.cols]
        self.to_datetime = to_datetime

    def __iter__(self):
        return self.__next__()

    def __next__(self):
        for n in self.ws.iter_rows(min_row=2, max_col=self.ws.max_column, max_row=self.ws.max_row,values_only=True):
            row = dict(zip(self.cols,n))
            if self.to_datetime:
                for k in row:
                    if k in self.to_datetime:
                        row[k] = parse(row[k])
            yield row

class source_csv(Source):
    def __init__(self,file,fieldnames=None,arraysize=300000,mapping ={},to_datetime = [],restkey = None,restval = None,dialect = None,delimiter=',',encoding = 'utf-8',*args, **kwds):
        '''读取csv文件 获取可迭代对象
        
        Parameters
        ------------
        file:str
            文件路径
        fieldnames:list
            指定字段
        arraysize:int
            数据交换行数 默认30000行 根据内存合理配置
        mapping:dict
            数据转换映射 {'a':int,'b':str,'c':func}
        to_datetime:list
            格式化为datetime数据类型
        delimiter:str
            字段分隔符号
        dialect:str
            字段分隔符
        '''
        if file.split('.')[-1] in ['GZ','gz']:
            self.file_open = gzip.open
        else:
            self.file_open = open
        self.csv_file = self.file_open(file, 'rt',encoding=encoding)
        self.csv_reader = csv.DictReader(self.csv_file,fieldnames = fieldnames,restkey=restkey,restval = restval,dialect = dialect,delimiter=delimiter,*args, **kwds)
        self.to_datetime = to_datetime
        self.data = queue.Queue(maxsize=arraysize)
        self.__obtain_thread = threading.Thread(target=self.__data_obtain)
        self.first_status = True
        self.last_status = False

    @property
    def fieldnames(self):
        return self.csv_reader.fieldnames

    @fieldnames.setter
    def fieldnames(self,fieldnames):
        self.csv_reader.fieldnames = fieldnames

    def __data_obtain(self):
        while 1:
            try:
                row = self.csv_reader.__next__()
                if self.to_datetime:
                    for k in row:
                        if k in self.to_datetime:
                            row[k] = parse(row[k])
                self.data.put(row)
            except StopIteration:
                self.last_status = True
                if getattr(self.csv_file,'close'):
                    self.csv_file.close()
                return 

            except Exception as e:
                self.last_status = True
                if getattr(self.csv_file,'close'):
                    self.csv_file.close()
                raise e

    def __iter__(self):
       return self

    def __next__(self):
        while 1:
            if self.first_status:
                self.__obtain_thread.start()
                self.first_status = False 

            if not self.last_status:
                return self.data.get()
            else:
                if self.data.empty():
                    raise StopIteration 
                else:
                    return self.data.get()

class source_es(object):
    def __init__(self,conn,index,**kwargs):
        '''elasticsearchc 数据查询
        conn:elasticsearch.client.Elasticsearch
        index:
            索引名称
        term:dict
            精确匹配
            {field:value}
        match:dict
            全文匹配
            {field:value}
        range:dict
            区间匹配 {num_field:{'lt':2}}
            lt:小于
            lte:小于或者等于
            gt:大于
            gte:大于等于
        fuzzy:dict
            模糊匹配
            {field:value}
        match_phrase:dict
            短语匹配
        match_phrase_prefix:dict
            匹配前缀
        multi_match:
            {'query':...,'fields': [...]}
        '''
        from elasticsearch_dsl import Search,Q
        
        self._s = Search(using=conn,index = index).query('bool',must=[Q(k.lower(),**v) for k,v in kwargs.items()])
    
    def __iter__(self):
        return self.__next__()
    
    def __next__(self):
        for row in self._s:
            yield row.to_dict()


