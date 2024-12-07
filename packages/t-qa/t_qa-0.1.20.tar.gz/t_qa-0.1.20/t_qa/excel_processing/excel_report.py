"""Module for excel class."""
from openpyxl.workbook import Workbook


class ExcelReport:
    """Excel class represent local excel file."""

    def __init__(self, local_file_path: str) -> None:
        """Init for excel class."""
        self.local_file_path = local_file_path
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)
        self.workbook.create_sheet("Report")
        self.worksheet = self.workbook["Report"]

    def write_cells_to_excel(self, all_cells: list) -> None:
        """Write the cells to the Excel file."""
        for cell in all_cells:
            self.worksheet.cell(row=cell.row, column=cell.column_number, value=cell.value)
        self.workbook.save(self.local_file_path)

    def write_header_to_excel(self, headers: list) -> None:
        """Write the header to the Excel file."""
        for cell in headers:
            self.worksheet.cell(row=cell.row, column=cell.column_number, value=cell.value)
