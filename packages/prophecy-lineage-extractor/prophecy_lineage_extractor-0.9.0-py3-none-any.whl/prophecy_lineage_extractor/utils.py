import logging
import logging
import os
import smtplib
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from urllib.parse import urlparse, urlunparse

import pandas as pd
import sqlparse
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook

from prophecy_lineage_extractor import constants
from prophecy_lineage_extractor.constants import (
    PIPELINE_ID,
)
from prophecy_lineage_extractor.constants import PROPHECY_URL


def get_prophecy_name(id):
    return id.split("/")[2]


def get_output_path(fmt="xlsx"):
    pipeline_name = get_prophecy_name(safe_env_variable(PIPELINE_ID))
    return Path(__file__).parent.parent / "output" / f"lineage_{pipeline_name}.{fmt}"


def delete_file(output_path):
    if output_path.exists():
        logging.info(f"Removing file{output_path} ")
        output_path.unlink()
    else:
        logging.info(f"file {output_path} doesn't exist, nothing to delete")


def _remove_nulls(df, columns):
    for column in columns:
        df[column] = df[column].fillna("")
        df[column] = df[column].replace("None", "")
    return df


def append_to_csv(df, csv_path):
    # Check if the DataFrame is empty
    if df.empty:
        logging.info("CSV_WRITER: Received an empty DataFrame, nothing to write.")
        return
    else:
        logging.info(f"CSV_WRITER: Appending data to {csv_path}")
    get_output_path().parent.mkdir(parents=True, exist_ok=True)
    # Append data to a CSV file, creating it if it doesn't exist
    with open(csv_path, "a", newline="") as f:
        df.to_csv(f, header=f.tell() == 0, index=False)


def pandas_configs():
    # Set options to display the full DataFrame
    pd.set_option("display.max_rows", None)  # Show all rows
    pd.set_option("display.max_columns", None)  # Show all columns
    pd.set_option(
        "display.max_colwidth", None
    )  # Show full column content without truncation
    pd.set_option("display.width", 1000)  # Set display width to avoid line wrapping


def convert_csv_to_excel(csv_path: Path, output_path: Path):
    # check if excel file already exists
    delete_file(get_output_path())
    get_output_path().parent.mkdir(parents=True, exist_ok=True)

    pandas_configs()

    df = pd.read_csv(csv_path)
    # Clean nulls in specific columns
    # df = _remove_nulls(df, ['upstream_transformation', 'downstream_transformation'])
    df = _remove_nulls(df, [constants.UPSTREAM_TRANSFORMATION_COL])

    # Define a function to concatenate transformations
    def _concatenate_transformations(series):
        return "\n\n".join(series)  # Concatenate

    df = df.groupby(
        [constants.DB_COL, constants.TABLE_COL, constants.COLNAME_COL], as_index=False
    ).agg({constants.UPSTREAM_TRANSFORMATION_COL: _concatenate_transformations})

    # Create a new Excel workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Convert DataFrame to rows and write to the Excel sheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(row)
        if r_idx == 1:  # Apply header styles
            for cell in ws[r_idx]:
                cell.fill = PatternFill(
                    start_color="FFDD99", end_color="FFDD99", fill_type="solid"
                )
                cell.font = Font(bold=True)

    EXCEL_DATABASE_COL = "A"
    EXCEL_TABLE_COL = "B"
    EXCEL_COLNAME_COL = "C"
    EXCEL_UPSTREAM_COL = "D"
    # DOWNSTREAM_COL = "D"
    # Define custom widths for the columns
    column_widths = {
        EXCEL_DATABASE_COL: 30,  # Width for 'database'
        EXCEL_TABLE_COL: 40,  # Width for 'table'
        EXCEL_COLNAME_COL: 30,  # Width for 'column_name'
        EXCEL_UPSTREAM_COL: 80,  # Width for 'upstream_transformation'
    }
    # Set the column widths
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Apply styles to headers
    header_fill = PatternFill(
        start_color="FFDD99", end_color="FFDD99", fill_type="solid"
    )  # Light orange for header
    for cell in ws[1]:  # First row as header
        cell.fill = header_fill
        cell.font = Font(bold=True)

    def _align_excel_column(col):
        for _cell in ws[col]:  # Column D corresponds to 'upstream_transformation'
            _cell.alignment = Alignment(wrapText=True, vertical="center")

    for key in column_widths.keys():
        _align_excel_column(key)

    # Save the workbook
    wb.save(output_path)
    logging.info(f"Success: Excel file saved to {output_path}")


# def save_excel_file(df: pd.DataFrame, output_path: Path):
#     # Check if DataFrame is empty
#     if df.empty:
#         logging.info("Empty DataFrame, nothing to write")
#         return
#     get_output_path().parent.mkdir(parents=True, exist_ok=True)
#     # Set options to display the full DataFrame
#     pd.set_option("display.max_rows", None)  # Show all rows
#     pd.set_option("display.max_columns", None)  # Show all columns
#     pd.set_option(
#         "display.max_colwidth", None
#     )  # Show full column content without truncation
#     pd.set_option("display.width", 1000)  # Set display width to avoid line wrapping
#     # logging.info(df)
#
#     # Clean nulls in specific columns
#     # df = _remove_nulls(df, ['upstream_transformation', 'downstream_transformation'])
#     df = _remove_nulls(df, ["upstream_transformation"])
#
#     # Check if the file already exists
#     if output_path.exists():
#         logging.info(f"Appending to existing file: {output_path}")
#         wb = load_workbook(output_path)
#         ws = wb.active
#         logging.debug(f"Number of rows already exists: {ws.max_row}")
#         start_row = ws.max_row + 1  # Start appending from the next row
#     else:
#         # Create a new workbook if the file doesn't exist
#         logging.info(f"Creating new file: {output_path}")
#         df.to_excel(output_path, index=False)  # Save initial data for headers
#         wb = load_workbook(output_path)
#         ws = wb.active
#         start_row = 2  # Starting row for new data (after headers)
#
#     # Append DataFrame rows to the worksheet
#     for row_data in df.itertuples(index=False, name=None):
#         ws.append(row_data)
#
#     # Apply styles to headers
#     header_fill = PatternFill(
#         start_color="FFDD99", end_color="FFDD99", fill_type="solid"
#     )  # Light orange for header
#     for cell in ws[1]:  # First row as header
#         cell.fill = header_fill
#         cell.font = Font(bold=True)
#
#     # Set specific column widths
#
#     DATABASE_COL = "A"
#     TABLE_COL = "B"
#     COLNAME_COL = "C"
#     UPSTREAM_COL = "D"
#     # DOWNSTREAM_COL = "D"
#     ws.column_dimensions[DATABASE_COL].width = 30  # Column name
#     ws.column_dimensions[TABLE_COL].width = 50  # Column name
#     ws.column_dimensions[COLNAME_COL].width = 30  # Column name
#     ws.column_dimensions[UPSTREAM_COL].width = 80  # Upstream transformation
#     # ws.column_dimensions[DOWNSTREAM_COL].width = 60  # Downstream transformation
#
#     # Process rows to merge "Name", "Type", and "Nullable" columns and concatenate "Upstream" and "Downstream"
#     current_row = start_row
#     while current_row <= ws.max_row:
#         database_col_name = ws[f"{DATABASE_COL}{current_row}"].value
#         table_col_name = ws[f"{TABLE_COL}{current_row}"].value
#         column_name_value = ws[f"{COLNAME_COL}{current_row}"].value
#         next_row = current_row + 1
#
#         # Initialize concatenated values for "Upstream" and "Downstream"
#         upstream_concat = ws[f"{UPSTREAM_COL}{current_row}"].value or ""
#         # downstream_concat = ws[f"{DOWNSTREAM_COL}{current_row}"].value or ""
#
#         # Check if the next rows have the same ID value
#         while (
#                 next_row <= ws.max_row
#                 and ws[f"{DATABASE_COL}{next_row}"].value == database_col_name
#                 and ws[f"{TABLE_COL}{next_row}"].value == table_col_name
#                 and ws[f"{COLNAME_COL}{next_row}"].value == column_name_value
#         ):
#             # Concatenate "Upstream" and "Downstream" values
#             if ws[f"{UPSTREAM_COL}{next_row}"].value:
#                 upstream_concat += f"\n\n{ws[f'{UPSTREAM_COL}{next_row}'].value}"
#             # if ws[f"{DOWNSTREAM_COL}{next_row}"].value:
#             #     downstream_concat += f"\n{ws[f'{DOWNSTREAM_COL}{next_row}'].value}"
#             next_row += 1
#
#         # Update "Upstream" and "Downstream" columns with concatenated values
#         ws[f"{UPSTREAM_COL}{current_row}"].value = upstream_concat
#         # ws[f"{DOWNSTREAM_COL}{current_row}"].value = downstream_concat
#         ws[f"{UPSTREAM_COL}{current_row}"].alignment = Alignment(
#             wrap_text=True, vertical="center"
#         )
#         # ws[f"{DOWNSTREAM_COL}{current_row}"].alignment = Alignment(wrap_text=True, vertical="center")
#
#         # Merge cells if there are multiple rows with the same "Name"
#         if next_row - current_row > 1:
#             ws.merge_cells(
#                 start_row=current_row,
#                 start_column=1,
#                 end_row=next_row - 1,
#                 end_column=1,
#             )  # Merge "Name" column
#             ws.merge_cells(
#                 start_row=current_row,
#                 start_column=2,
#                 end_row=next_row - 1,
#                 end_column=2,
#             )  # Merge "Type" column
#             ws.merge_cells(
#                 start_row=current_row,
#                 start_column=3,
#                 end_row=next_row - 1,
#                 end_column=3,
#             )  # Merge "Nullable" column
#
#         # Move to the next distinct value
#         current_row = next_row
#
#     # Save changes
#     wb.save(output_path)
#     logging.info(
#         f"Excel file with merged cells and concatenated transformations saved to {output_path}"
#     )


def get_ws_url():
    prophecy_url = safe_env_variable(PROPHECY_URL)
    try:
        # Parse the URL
        parsed_url = urlparse(prophecy_url)

        # # Ensure the URL uses HTTPS
        # if parsed_url.scheme != "https":
        #     raise ValueError("Invalid URL. Must start with 'https://'.")

        # Remove 'www.' from the netloc (hostname)
        netloc = parsed_url.netloc.replace("www.", "")

        # Create the WebSocket URL

        # Create the WebSocket URL
        websocket_url = parsed_url._replace(
            scheme="wss", netloc=netloc, path="/api/lineage/ws"
        )

        # Return the reconstructed URL without trailing slashes
        return urlunparse(websocket_url).rstrip("/")
    except Exception as e:
        raise ValueError(f"Error processing URL: {e}")


def get_graphql_url():
    prophecy_url = safe_env_variable(PROPHECY_URL)

    try:
        parsed_url = urlparse(prophecy_url)
        # # Ensure the URL uses HTTPS
        # if parsed_url.scheme not in ["https", "http"]:
        #     raise ValueError("Invalid URL. Must start with 'https://' or 'http://'.")

        # Remove 'www.' from the netloc (hostname)
        netloc = parsed_url.netloc.replace("www.", "")
        # Append '/api/md/graphql' to the path
        path = parsed_url.path.rstrip("/") + "/api/md/graphql"
        # Create the modified URL
        modified_url = parsed_url._replace(netloc=netloc, path=path)
        # Return the reconstructed URL
        return urlunparse(modified_url)

    except Exception as e:
        raise ValueError(f"Error processing URL: {e}")


def safe_env_variable(var_name):
    if var_name not in os.environ:
        logging.error(
            f"[ERROR]: Environment variable '{var_name}' is not set, Please set this value to continue."
        )
        raise Exception(f"Environment variable '{var_name}' is not set")
    return os.environ[var_name]  # Optional: return the value if needed.


def send_excel_email(file_path: Path):
    # Get SMTP credentials and email info from environment variables
    smtp_host = safe_env_variable("SMTP_HOST")
    smtp_port = int(safe_env_variable("SMTP_PORT"))  # with default values
    smtp_username = safe_env_variable("SMTP_USERNAME")
    smtp_password = safe_env_variable("SMTP_PASSWORD")
    receiver_email = safe_env_variable("RECEIVER_EMAIL")

    if not all([smtp_host, smtp_port, smtp_username, smtp_password, receiver_email]):
        raise ValueError("Missing required environment variables for SMTP or email.")

    # Create email message
    msg = MIMEMultipart()
    msg["From"] = smtp_username
    msg["To"] = receiver_email
    msg["Subject"] = (
        f"Prophecy Lineage report for Pipeline: {get_prophecy_name(safe_env_variable(PIPELINE_ID))}"
    )

    # Email body
    body = (
        f"Dear user,\n\tPlease find the attached Prophecy Lineage Excel report for "
        f"Pipeline Id: {safe_env_variable(PIPELINE_ID)} \n\nThanks and regards,\n\tProphecy Team"
    )
    msg.attach(MIMEText(body, "plain"))

    # Attach Excel file
    attachment_name = file_path.name
    with open(file_path, "rb") as attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition", f"attachment; filename= {attachment_name}"
        )
        msg.attach(part)

    # Send email via SMTP
    try:
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(smtp_username, smtp_password)
            server.send_message(msg)
            logging.info(f"Email sent successfully to {receiver_email}")
    except Exception as e:
        logging.error(f"Failed to send email: {str(e)}")
        raise e


def _is_sql(expr):
    sql_keywords = {
        "SELECT",
        "INSERT",
        "UPDATE",
        "DELETE",
        "CREATE",
        "DROP",
        "ALTER",
        "FROM",
        "WHERE",
        "JOIN",
    }
    python_keywords = {
        ".",
        "def",
        "lambda",
        "import",
        "class",
        "return",
        "self",
        "=",
        "(",
        ")",
    }
    scala_keywords = {
        "val",
        "var",
        "def",
        "object",
        "class",
        "case",
        "match",
        "=>",
        ":",
        "implicit",
    }

    if any(py_kw in expr for py_kw in python_keywords) or any(
        scala_kw in expr for scala_kw in scala_keywords
    ):
        return False
    words = expr.strip().upper().split()
    return any(word in sql_keywords for word in words)


def format_sql(expr):
    try:
        # if _is_sql(expr):
        formatted_query = sqlparse.format(expr, reindent=True, keyword_case="upper")
        logging.info("query formatted.")
        return formatted_query
    # else:
    #     return expr
    except Exception as e:
        logging.error(
            f"Error occurred while formatting sql expression, returning original: \n {expr}\n error: {e}"
        )
        return expr
