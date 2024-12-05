import json
import logging
import os
import smtplib
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from urllib.parse import urlparse, urlunparse

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

from prophecy_lineage_extractor.constants import PROJECT_ID, PIPELINE_ID
from prophecy_lineage_extractor.constants import PROPHECY_URL, PROPHECY_PAT


def get_output_path():
    pipeline_name = safe_env_variable(PIPELINE_ID).split("/")[2]
    return Path(__file__).parent.parent / "output" / f"lineage_{pipeline_name}.xlsx"


def delete_file(output_path):
    if output_path.exists():
        logging.warning(f"Removing file{output_path} ")
        output_path.unlink()
    else:
        logging.warning(f"file {output_path} doesn't exist, nothing to delete")


def _remove_nulls(df, columns):
    for column in columns:
        df[column] = df[column].fillna("")
        df[column] = df[column].replace("None", "")
    return df


def save_excel_file(df: pd.DataFrame, output_path: Path):
    # Check if DataFrame is empty
    if df.empty:
        logging.warning("Empty DataFrame, nothing to write")
        return
    get_output_path().parent.mkdir(parents=True, exist_ok=True)
    # Set options to display the full DataFrame
    pd.set_option("display.max_rows", None)  # Show all rows
    pd.set_option("display.max_columns", None)  # Show all columns
    pd.set_option(
        "display.max_colwidth", None
    )  # Show full column content without truncation
    pd.set_option("display.width", 1000)  # Set display width to avoid line wrapping
    logging.warning(df)

    # Clean nulls in specific columns
    # df = _remove_nulls(df, ['upstream_transformation', 'downstream_transformation'])
    df = _remove_nulls(df, ["upstream_transformation"])

    # Check if the file already exists
    if output_path.exists():
        logging.warning(f"Appending to existing file: {output_path}")
        wb = load_workbook(output_path)
        ws = wb.active
        logging.debug(f"Number of rows already exists: {ws.max_row}")
        start_row = ws.max_row + 1  # Start appending from the next row
    else:
        # Create a new workbook if the file doesn't exist
        logging.warning(f"Creating new file: {output_path}")
        df.to_excel(output_path, index=False)  # Save initial data for headers
        wb = load_workbook(output_path)
        ws = wb.active
        start_row = 2  # Starting row for new data (after headers)

    # Append DataFrame rows to the worksheet
    for row_data in df.itertuples(index=False, name=None):
        ws.append(row_data)

    # Apply styles to headers
    header_fill = PatternFill(
        start_color="FFDD99", end_color="FFDD99", fill_type="solid"
    )  # Light orange for header
    for cell in ws[1]:  # First row as header
        cell.fill = header_fill
        cell.font = Font(bold=True)

    # Set specific column widths

    DATABASE_COL = "A"
    TABLE_COL = "B"
    COLNAME_COL = "C"
    UPSTREAM_COL = "D"
    # DOWNSTREAM_COL = "D"
    ws.column_dimensions[DATABASE_COL].width = 30  # Column name
    ws.column_dimensions[TABLE_COL].width = 50  # Column name
    ws.column_dimensions[COLNAME_COL].width = 30  # Column name
    ws.column_dimensions[UPSTREAM_COL].width = 80  # Upstream transformation
    # ws.column_dimensions[DOWNSTREAM_COL].width = 60  # Downstream transformation

    # Process rows to merge "Name", "Type", and "Nullable" columns and concatenate "Upstream" and "Downstream"
    current_row = start_row
    while current_row <= ws.max_row:
        database_col_name = ws[f"{DATABASE_COL}{current_row}"].value
        table_col_name = ws[f"{TABLE_COL}{current_row}"].value
        column_name_value = ws[f"{COLNAME_COL}{current_row}"].value
        next_row = current_row + 1

        # Initialize concatenated values for "Upstream" and "Downstream"
        upstream_concat = ws[f"{UPSTREAM_COL}{current_row}"].value or ""
        # downstream_concat = ws[f"{DOWNSTREAM_COL}{current_row}"].value or ""

        # Check if the next rows have the same "Name" value
        while (
            next_row <= ws.max_row
            and ws[f"{DATABASE_COL}{next_row}"].value == database_col_name
            and ws[f"{TABLE_COL}{next_row}"].value == table_col_name
            and ws[f"{COLNAME_COL}{next_row}"].value == column_name_value
        ):
            # Concatenate "Upstream" and "Downstream" values
            if ws[f"{UPSTREAM_COL}{next_row}"].value:
                upstream_concat += f"\n{ws[f'{UPSTREAM_COL}{next_row}'].value}"
            # if ws[f"{DOWNSTREAM_COL}{next_row}"].value:
            #     downstream_concat += f"\n{ws[f'{DOWNSTREAM_COL}{next_row}'].value}"
            next_row += 1

        # Update "Upstream" and "Downstream" columns with concatenated values
        ws[f"{UPSTREAM_COL}{current_row}"].value = upstream_concat
        # ws[f"{DOWNSTREAM_COL}{current_row}"].value = downstream_concat
        ws[f"{UPSTREAM_COL}{current_row}"].alignment = Alignment(
            wrap_text=True, vertical="center"
        )
        # ws[f"{DOWNSTREAM_COL}{current_row}"].alignment = Alignment(wrap_text=True, vertical="center")

        # Merge cells if there are multiple rows with the same "Name"
        if next_row - current_row > 1:
            ws.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=next_row - 1,
                end_column=1,
            )  # Merge "Name" column
            ws.merge_cells(
                start_row=current_row,
                start_column=2,
                end_row=next_row - 1,
                end_column=2,
            )  # Merge "Type" column
            ws.merge_cells(
                start_row=current_row,
                start_column=3,
                end_row=next_row - 1,
                end_column=3,
            )  # Merge "Nullable" column

        # Move to the next distinct "Name" value
        current_row = next_row

    # Save changes
    wb.save(output_path)
    logging.warning(
        f"Excel file with merged cells and concatenated transformations saved to {output_path}"
    )


def get_ws_url():
    prophecy_url = safe_env_variable(PROPHECY_URL)
    try:
        # Parse the URL
        parsed_url = urlparse(prophecy_url)

        # # Ensure the URL uses HTTPS
        # if parsed_url.scheme != "https":
        #     raise ValueError("Invalid URL. Must start with 'https://'.")

        # Remove 'www.' from the netloc (hostname)
        netloc = parsed_url.netloc.replace("www.", "")

        # Create the WebSocket URL

        # Create the WebSocket URL
        websocket_url = parsed_url._replace(
            scheme="wss", netloc=netloc, path="/api/lineage/ws"
        )

        # Return the reconstructed URL without trailing slashes
        return urlunparse(websocket_url).rstrip("/")
    except Exception as e:
        raise ValueError(f"Error processing URL: {e}")


def get_graphql_url():
    prophecy_url = safe_env_variable(PROPHECY_URL)

    try:
        parsed_url = urlparse(prophecy_url)
        # # Ensure the URL uses HTTPS
        # if parsed_url.scheme not in ["https", "http"]:
        #     raise ValueError("Invalid URL. Must start with 'https://' or 'http://'.")

        # Remove 'www.' from the netloc (hostname)
        netloc = parsed_url.netloc.replace("www.", "")
        # Append '/api/md/graphql' to the path
        path = parsed_url.path.rstrip("/") + "/api/md/graphql"
        # Create the modified URL
        modified_url = parsed_url._replace(netloc=netloc, path=path)
        # Return the reconstructed URL
        return urlunparse(modified_url)

    except Exception as e:
        raise ValueError(f"Error processing URL: {e}")


def safe_env_variable(var_name):
    if var_name not in os.environ:
        logging.error(
            f"[ERROR]: Environment variable '{var_name}' is not set, Please set this value to continue."
        )
        raise Exception(f"Environment variable '{var_name}' is not set")
    return os.environ[var_name]  # Optional: return the value if needed.


def send_excel_email(file_path: Path):
    # Get SMTP credentials and email info from environment variables
    smtp_host = safe_env_variable("SMTP_HOST")
    smtp_port = int(safe_env_variable("SMTP_PORT"))  # with default values
    smtp_username = safe_env_variable("SMTP_USERNAME")
    smtp_password = safe_env_variable("SMTP_PASSWORD")
    receiver_email = safe_env_variable("RECEIVER_EMAIL")

    if not all([smtp_host, smtp_port, smtp_username, smtp_password, receiver_email]):
        raise ValueError("Missing required environment variables for SMTP or email.")

    # Create email message
    msg = MIMEMultipart()
    msg["From"] = smtp_username
    msg["To"] = receiver_email
    msg["Subject"] = (
        f"Prophecy Lineage report for Pipeline: {safe_env_variable(PIPELINE_ID).split('/')[2]}"
    )

    # Email body
    body = (
        f"Dear user,\n\tPlease find the attached Prophecy Lineage Excel report for "
        f"Pipeline Id: {safe_env_variable(PIPELINE_ID)}<\n\nThanks and regards,\n\tProphecy Team"
    )
    msg.attach(MIMEText(body, "plain"))

    # Attach Excel file
    attachment_name = file_path.name
    with open(file_path, "rb") as attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition", f"attachment; filename= {attachment_name}"
        )
        msg.attach(part)

    # Send email via SMTP
    try:
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(smtp_username, smtp_password)
            server.send_message(msg)
            logging.info(f"Email sent successfully to {receiver_email}")
    except Exception as e:
        logging.error(f"Failed to send email: {str(e)}")
        raise e
